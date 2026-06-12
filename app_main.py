import sys
import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from pathlib import Path
from datetime import datetime
import csv
import re
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict
import threading

# Import blockchain scripts later (inside run() function after API keys are set)

# ========================
# CONFIG FILE MANAGEMENT
# ========================
CONFIG_FILE = Path.home() / ".crypto_analyzer_config.json"

def load_config():
    """Load API keys from config file."""
    if CONFIG_FILE.exists():
        try:
            with open(CONFIG_FILE, 'r') as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_config(config):
    """Save API keys to config file."""
    try:
        with open(CONFIG_FILE, 'w') as f:
            json.dump(config, f)
        return True
    except:
        return False

# ========================
# EXCEL PROCESSING
# ========================
def normalise_columns(columns):
    out = []
    for c in columns:
        c = str(c).strip().lower()
        c = re.sub(r'[\s/()]+', '_', c)
        c = re.sub(r'_+', '_', c).strip('_')
        out.append(c)
    return out

def detect_header_and_clean(df):
    header_row_idx = 0
    header = df.iloc[header_row_idx]
    body = df.iloc[header_row_idx + 1:].reset_index(drop=True)
    body.columns = normalise_columns(header)

    seen = set()
    keep_mask = []
    for c in body.columns:
        if c == '' or c in seen:
            keep_mask.append(False)
        else:
            seen.add(c)
            keep_mask.append(True)
    body = body.loc[:, keep_mask]

    if 'balance_date' in body.columns:
        body['timestamp'] = pd.to_datetime(
            body['balance_date'].astype(str).str.strip(),
            errors='coerce',
            utc=True,
            dayfirst=True
        )

    return body.dropna(how="all")

def read_folder(folder_path):
    folder = Path(folder_path)
    files = sorted(folder.glob('*.xlsx'))
    if not files:
        raise FileNotFoundError(f"No XLSX files in {folder}")

    frames = []
    for f in files:
        print(f"Reading {f.name}...")
        wb = load_workbook(f, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        raw = pd.DataFrame(rows)
        df = detect_header_and_clean(raw)
        if df.empty:
            print(f"  skip: no header row found in {f.name}")
            continue
        df['source_file'] = f.name
        frames.append(df)
        print(f"  {len(df):,} rows")

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)
    print(f"\nCombined: {len(combined):,} rows from {len(frames)} file(s)")
    return combined

def run(folder_path, ondate, log_callback=None):
    """Run analysis - process wallets and generate report."""
    # Import blockchain scripts HERE after API keys are set
    from solana_alchemy import solana_wallet_value_on_date
    from eth_alchemy import eth_snapshot_with_usd
    
    def log(msg):
        print(msg)
        if log_callback:
            log_callback(msg)

    log(f"Loading files from {folder_path}...")
    df = read_folder(folder_path)

    solana = []
    ethereum = []
    other = []

    for idx, row in df.iterrows():
        if row['ticker'] and ('SOL' in row['ticker']):
            solana.append(row['wallet'])
        if row['ticker'] and ('ETH' in row['ticker']):
            ethereum.append(row['wallet'])
        if row['ticker'] and all(x not in row['ticker'] for x in ['ETH', 'SOL']):
            other.append((row['ticker'], row['wallet']))

    log(f"Found {len(solana)} Solana wallets, {len(ethereum)} Ethereum wallets, {len(other)} other")

    data = defaultdict(list)
    v = ['date','wallet','ticker','main_amt','main_usd_price','main_usd_value',
         'usdc_usd_value','usdt_usd_value','usx_usd_value','jitosol_usd_value','weth_usd_val']
    for i in v:
        data[i] = []

    # SOLANA
    for i, wallet in enumerate(solana):
        try:
            log(f"[{i+1}/{len(solana)}] Processing Solana wallet: {wallet[:16]}...")
            (sol, sol_usd_price, sol_usd_value, usdc_usd_value, usdt_usd_value,
             usx_usd_value, jitosol_usd_value) = solana_wallet_value_on_date(wallet, ondate, debug=False)

            data['date'].append(ondate)
            data['wallet'].append(wallet)
            data['ticker'].append("SOL")
            data['main_amt'].append(sol)
            data['main_usd_price'].append(sol_usd_price)
            data['main_usd_value'].append(sol_usd_value)
            data['usdc_usd_value'].append(usdc_usd_value)
            data['usdt_usd_value'].append(usdt_usd_value)
            data['usx_usd_value'].append(usx_usd_value)
            data['jitosol_usd_value'].append(jitosol_usd_value)
            data['weth_usd_val'].append(None)
            log(f"  ✓ SOL: {sol:.4f}, USD Value: ${sol_usd_value:,.2f}")
        except Exception as e:
            log(f"  ✗ Error: {str(e)}")

    # ETHEREUM
    for i, wallet in enumerate(ethereum):
        try:
            log(f"[{i+1}/{len(ethereum)}] Processing Ethereum wallet: {wallet[:16]}...")
            (eth_amt, eth_usd_rate, eth_usd_value, usdc_usd_value, usdt_usd_value,
             weth_usd_value) = eth_snapshot_with_usd(wallet, ondate)

            data['date'].append(ondate)
            data['wallet'].append(wallet)
            data['ticker'].append("ETH")
            data['main_amt'].append(eth_amt)
            data['main_usd_price'].append(eth_usd_rate)
            data['main_usd_value'].append(eth_usd_value)
            data['usdc_usd_value'].append(usdc_usd_value)
            data['usdt_usd_value'].append(usdt_usd_value)
            data['usx_usd_value'].append(None)
            data['jitosol_usd_value'].append(None)
            data['weth_usd_val'].append(weth_usd_value)
            log(f"  ✓ ETH: {eth_amt:.4f}, USD Value: ${eth_usd_value:,.2f}")
        except Exception as e:
            log(f"  ✗ Error: {str(e)}")

    # OTHER
    for wallet_info in other:
        ticker, wallet = wallet_info
        log(f"Processing {ticker} wallet: {wallet[:16]}...")
        data['date'].append(ondate)
        data['wallet'].append(wallet)
        data['ticker'].append(ticker)
        data['main_amt'].append(None)
        data['main_usd_price'].append(None)
        data['main_usd_value'].append(None)
        data['usdc_usd_value'].append(None)
        data['usdt_usd_value'].append(None)
        data['usx_usd_value'].append(None)
        data['jitosol_usd_value'].append(None)
        data['weth_usd_val'].append(None)
        log(f"  ⚠ {ticker}: Not yet supported")

    data_onchain = pd.DataFrame(data)
    log(f"\n✓ Analysis complete! {len(data_onchain)} rows")
    return df, data_onchain

# ========================
# GUI APPLICATION
# ========================
class CryptoAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Crypto Wallet Analyzer")
        self.root.geometry("700x600")
        self.config = load_config()
        self.is_running = False

        self.setup_ui()
        self.check_api_keys()

    def setup_ui(self):
        # ---- HEADER ----
        header = ttk.Frame(self.root)
        header.pack(fill='x', padx=20, pady=15)
        
        ttk.Label(header, text="💰 Crypto Wallet Analyzer", font=('Arial', 16, 'bold')).pack()
        ttk.Label(header, text="Analyze SOL, ETH, BTC wallets on any date", 
                  font=('Arial', 9), foreground='gray').pack()

        # ---- API KEYS SECTION ----
        api_frame = ttk.LabelFrame(self.root, text="🔑 API Configuration", padding=10)
        api_frame.pack(fill='x', padx=20, pady=10)

        ttk.Button(api_frame, text="⚙️  Set API Keys", command=self.show_api_dialog).pack(fill='x', pady=5)
        self.api_status = ttk.Label(api_frame, text="Status: Not configured", foreground='red')
        self.api_status.pack(anchor='w')

        # ---- FOLDER SECTION ----
        folder_frame = ttk.LabelFrame(self.root, text="📁 Select Input Folder", padding=10)
        folder_frame.pack(fill='x', padx=20, pady=10)

        self.folder_var = tk.StringVar(value="Click to select...")
        ttk.Button(folder_frame, text="📂 Browse", command=self.select_folder).pack(fill='x', pady=5)
        self.folder_label = ttk.Label(folder_frame, text=self.folder_var.get(), 
                                      wraplength=600, foreground='blue')
        self.folder_label.pack(anchor='w', pady=5)

        def update_label(*args):
            self.folder_label.config(text=self.folder_var.get())
        self.folder_var.trace('w', update_label)

        # ---- DATE SECTION ----
        date_frame = ttk.LabelFrame(self.root, text="📅 Analysis Date", padding=10)
        date_frame.pack(fill='x', padx=20, pady=10)

        ttk.Label(date_frame, text="Date (YYYY-MM-DD):").pack(anchor='w')
        self.date_entry = ttk.Entry(date_frame, width=20)
        self.date_entry.pack(anchor='w', pady=5)
        self.date_entry.insert(0, datetime.now().strftime("%Y-%m-%d"))

        # ---- RUN BUTTON ----
        self.run_button = ttk.Button(self.root, text="🚀 Run Analysis", 
                                      command=self.run_analysis_threaded)
        self.run_button.pack(fill='x', padx=20, pady=10)

        # ---- LOG OUTPUT ----
        log_frame = ttk.LabelFrame(self.root, text="📊 Status Log", padding=5)
        log_frame.pack(fill='both', expand=True, padx=20, pady=10)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, width=80, 
                                                   font=('Courier', 9))
        self.log_text.pack(fill='both', expand=True)

        # ---- FOOTER ----
        ttk.Label(self.root, text="✓ Place Excel files in selected folder. Output saved to same location.",
                  font=('Arial', 8), foreground='gray').pack(side='bottom', pady=10)

    def check_api_keys(self):
        """Check if API keys are configured."""
        alchemy = self.config.get('ALCHEMY_KEY', '')
        etherscan = self.config.get('ETHERSCAN_KEY', '')
        
        if alchemy and etherscan:
            self.api_status.config(text=f"✓ API keys configured", foreground='green')
        else:
            self.api_status.config(text=f"⚠ Missing: {'ALCHEMY_KEY ' if not alchemy else ''}"
                                        f"{'ETHERSCAN_KEY' if not etherscan else ''}", 
                                   foreground='orange')

    def show_api_dialog(self):
        """Show dialog to enter API keys."""
        dialog = tk.Toplevel(self.root)
        dialog.title("API Configuration")
        dialog.geometry("500x300")
        dialog.resizable(False, False)

        ttk.Label(dialog, text="Enter your API keys below:", font=('Arial', 11, 'bold')).pack(pady=15)

        ttk.Label(dialog, text="Alchemy Key:").pack(anchor='w', padx=20)
        alchemy_var = tk.StringVar(value=self.config.get('ALCHEMY_KEY', ''))
        ttk.Entry(dialog, textvariable=alchemy_var, width=50).pack(padx=20, pady=5, fill='x')
        ttk.Label(dialog, text="Get from: https://dashboard.alchemy.com", 
                  font=('Arial', 8), foreground='blue').pack(anchor='w', padx=20)

        ttk.Label(dialog, text="Etherscan Key:").pack(anchor='w', padx=20, pady=(15, 0))
        etherscan_var = tk.StringVar(value=self.config.get('ETHERSCAN_KEY', ''))
        ttk.Entry(dialog, textvariable=etherscan_var, width=50).pack(padx=20, pady=5, fill='x')
        ttk.Label(dialog, text="Get from: https://etherscan.io/apis", 
                  font=('Arial', 8), foreground='blue').pack(anchor='w', padx=20)

        def save_keys():
            self.config['ALCHEMY_KEY'] = alchemy_var.get().strip()
            self.config['ETHERSCAN_KEY'] = etherscan_var.get().strip()
            
            if save_config(self.config):
                # Set environment variables
                os.environ['ALCHEMY_KEY'] = self.config['ALCHEMY_KEY']
                os.environ['ETHERSCAN_KEY'] = self.config['ETHERSCAN_KEY']
                messagebox.showinfo("Success", "API keys saved!")
                self.check_api_keys()
                dialog.destroy()
            else:
                messagebox.showerror("Error", "Failed to save keys")

        ttk.Button(dialog, text="💾 Save Keys", command=save_keys).pack(pady=15)

    def select_folder(self):
        folder = filedialog.askdirectory(title="Select folder with Excel files")
        if folder:
            self.folder_var.set(folder)

    def log(self, msg):
        self.log_text.insert('end', msg + '\n')
        self.log_text.see('end')
        self.root.update()

    def run_analysis_threaded(self):
        """Run analysis in a separate thread to keep GUI responsive."""
        if self.is_running:
            messagebox.showwarning("In Progress", "Analysis already running!")
            return

        folder_path = self.folder_var.get()
        date_str = self.date_entry.get()

        if "Click to select" in folder_path or not folder_path:
            messagebox.showerror("Error", "Please select a folder with Excel files")
            return

        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Error", "Invalid date. Use format: YYYY-MM-DD")
            return

        if not self.config.get('ALCHEMY_KEY') or not self.config.get('ETHERSCAN_KEY'):
            messagebox.showerror("Error", "Please configure API keys first")
            return

        # Set environment variables
        os.environ['ALCHEMY_KEY'] = self.config['ALCHEMY_KEY']
        os.environ['ETHERSCAN_KEY'] = self.config['ETHERSCAN_KEY']

        self.is_running = True
        self.run_button.config(state='disabled')
        self.log_text.delete(1.0, 'end')

        def thread_run():
            try:
                self.log(f"🔄 Starting analysis for {date_str}...\n")
                df, data_onchain = run(folder_path, date_str, log_callback=self.log)
                
                output_file = Path(folder_path) / f"analysis_{date_str}.csv"
                data_onchain.to_csv(output_file, index=False)
                
                self.log(f"\n✅ SUCCESS!\nFile saved to: {output_file}")
                self.log(f"Total wallets analyzed: {len(data_onchain)}")
                messagebox.showinfo("Success", f"Analysis complete!\n\nResults saved to:\n{output_file}")
            except Exception as e:
                self.log(f"\n❌ ERROR: {str(e)}")
                messagebox.showerror("Error", f"Analysis failed:\n{str(e)}")
            finally:
                self.is_running = False
                self.run_button.config(state='normal')

        import threading
        thread = threading.Thread(target=thread_run, daemon=True)
        thread.start()

# ========================
# MAIN
# ========================
if __name__ == "__main__":
    root = tk.Tk()
    app = CryptoAnalyzerApp(root)
    root.mainloop()